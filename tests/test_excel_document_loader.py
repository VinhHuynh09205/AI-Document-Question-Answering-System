from pathlib import Path

from openpyxl.comments import Comment
from openpyxl import Workbook

from app.services.document_loaders.excel_document_loader import ExcelDocumentLoader


def _write_sample_workbook(file_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Name", "Score", "Dept"])
    sheet.append(["Alice", 10, "A"])
    sheet.append(["Bob", 9, "B"])
    workbook.save(file_path)
    workbook.close()


def _write_workbook_with_title_row(file_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["試験結果表", "", "", "", "", "", "", "", "", ""])
    sheet.append(["No.", "受験番号", "氏名", "性別", "学科名", "日本語", "数学", "文学", "総計", "結果"])
    sheet.append([1, "KI2", "のび太 徳田", "男", "教育学科", 9.0, 8.5, 7.5, 34, "合格"])
    workbook.save(file_path)
    workbook.close()


def test_excel_loader_emits_sheet_summary_and_row_docs(tmp_path: Path) -> None:
    excel_path = tmp_path / "scores.xlsx"
    _write_sample_workbook(excel_path)

    loader = ExcelDocumentLoader()
    docs = loader.load(excel_path)

    assert len(docs) == 2

    sheet_doc = next(doc for doc in docs if doc.metadata.get("content_type") == "spreadsheet_sheet_summary")
    table_doc = next(doc for doc in docs if doc.metadata.get("content_type") == "spreadsheet_table")

    assert "File: scores.xlsx" in sheet_doc.page_content
    assert "Sheet: Sales" in sheet_doc.page_content
    assert sheet_doc.metadata.get("sheet_index") == 1
    assert sheet_doc.metadata.get("row_count") == 3

    assert table_doc.metadata.get("table_name") == "used_range_1"
    assert table_doc.metadata.get("range_address") == "A1:C3"
    assert table_doc.metadata.get("row_count") == 2
    assert table_doc.metadata.get("headers") == ["Name", "Score", "Dept"]
    structured_rows = table_doc.metadata.get("structured_rows")
    assert isinstance(structured_rows, list)
    assert len(structured_rows) == 2
    assert structured_rows[0]["row_number"] == 2
    assert structured_rows[0]["values"]["Name"] == "Alice"
    assert structured_rows[1]["values"]["Dept"] == "B"


def test_excel_loader_skips_title_row_and_uses_real_header_row(tmp_path: Path) -> None:
    excel_path = tmp_path / "exam_results.xlsx"
    _write_workbook_with_title_row(excel_path)

    loader = ExcelDocumentLoader()
    docs = loader.load(excel_path)

    sheet_doc = next(doc for doc in docs if doc.metadata.get("content_type") == "spreadsheet_sheet_summary")
    table_doc = next(doc for doc in docs if doc.metadata.get("content_type") == "spreadsheet_table")

    assert "Header Columns: No., 受験番号, 氏名, 性別, 学科名, 日本語, 数学, 文学, 総計, 結果" in sheet_doc.page_content
    assert table_doc.metadata.get("has_header") is True
    assert table_doc.metadata.get("header_row_number") == 2
    assert table_doc.metadata.get("row_count") == 1
    rows = table_doc.metadata.get("structured_rows")
    assert isinstance(rows, list)
    assert rows[0]["row_number"] == 3
    assert rows[0]["values"]["No."] == "1"
    assert rows[0]["values"]["総計"] == "34"
    assert rows[0]["values"]["結果"] == "合格"


def test_excel_loader_preserves_formula_comment_hyperlink_and_hidden_sheet(tmp_path: Path) -> None:
    excel_path = tmp_path / "rich.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["A", "B", "Total", "Reference"])
    sheet.append([10, 15, "=A2+B2", "Spec"])
    sheet["A2"].comment = Comment("Important input", "tester")
    sheet["D2"].hyperlink = "https://example.com/spec"
    sheet.merge_cells("A1:B1")

    hidden_sheet = workbook.create_sheet("HiddenStats")
    hidden_sheet.sheet_state = "hidden"
    hidden_sheet.append(["Metric", "Value"])
    hidden_sheet.append(["Count", 2])

    workbook.save(excel_path)
    workbook.close()

    loader = ExcelDocumentLoader()
    docs = loader.load(excel_path)

    hidden_summary = next(
        doc for doc in docs
        if doc.metadata.get("content_type") == "spreadsheet_sheet_summary"
        and doc.metadata.get("sheet_name") == "HiddenStats"
    )
    assert hidden_summary.metadata.get("sheet_hidden") is True

    data_table = next(
        doc for doc in docs
        if doc.metadata.get("content_type") == "spreadsheet_table"
        and doc.metadata.get("sheet_name") == "Data"
    )
    assert "A1:B1" in data_table.metadata.get("merged_ranges", [])

    rows = data_table.metadata.get("structured_rows")
    assert isinstance(rows, list)
    assert len(rows) >= 1
    first_row_cells = rows[0]["cells"]

    formula_cell = next(cell for cell in first_row_cells if cell.get("address") == "C2")
    assert formula_cell.get("formula") == "=A2+B2"

    comment_cell = next(cell for cell in first_row_cells if cell.get("address") == "A2")
    assert "Important input" in comment_cell.get("comment", "")

    hyperlink_cell = next(cell for cell in first_row_cells if cell.get("address") == "D2")
    assert hyperlink_cell.get("hyperlink") == "https://example.com/spec"


def test_excel_loader_merges_cached_formula_values_from_fixture() -> None:
    fixture_path = Path(__file__).resolve().parents[1] / "tmp" / "manual_test_files" / "Test xlsx.xlsx"

    loader = ExcelDocumentLoader()
    docs = loader.load(fixture_path)

    sheet1_table = next(
        doc for doc in docs
        if doc.metadata.get("content_type") == "spreadsheet_table"
        and doc.metadata.get("sheet_name") == "Sheet1"
    )

    rows = sheet1_table.metadata.get("structured_rows")
    assert isinstance(rows, list)

    first_candidate = next(row for row in rows if row.get("row_number") == 3)
    values = first_candidate.get("values")
    assert isinstance(values, dict)
    assert values["受験番号"] == "KI2"
    assert values["学科名"] == "教育学科"
    assert values["総計"] == "34"
    assert values["結果"] == "合格"

    cells = first_candidate.get("cells")
    assert isinstance(cells, list)

    total_cell = next(cell for cell in cells if cell.get("address") == "I3")
    assert total_cell.get("value") == "34"
    assert total_cell.get("formula") == "=IF(E3=$E$24,SUM(F3:H3),(F3*2)+G3+H3)"

    result_cell = next(cell for cell in cells if cell.get("address") == "J3")
    assert result_cell.get("value") == "合格"
    assert result_cell.get("formula") == '=IF(I3>=25,"合格","不合格")'
