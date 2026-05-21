from __future__ import annotations

import argparse
import csv
import json
import math
import re
import signal
import statistics
import time
import unicodedata
from collections import Counter, defaultdict, OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from langchain_core.documents import Document

from app.api.workspace import (
    _ensure_workspace_scope_indexed,
    _inject_document_mapping_into_question,
    _resolve_ask_routing,
)
from app.core.config import get_settings
from app.core.container import build_container


STOPWORDS = {
    "file", "slide", "sheet", "row", "rows", "page", "trang", "cau", "hoi",
    "noi", "dung", "trong", "tai", "lieu", "document", "chunk", "title",
    "layout", "reading", "order", "text", "table", "image", "data", "source",
    "metadata", "sample", "columns", "header", "headers", "index", "range",
    "this", "that", "with", "from", "and", "the", "cho", "cac", "mot",
    "nhung", "duoc", "khong", "co", "la", "ve", "vao", "cua", "cac",
    "nguoi", "dung", "thong", "tin", "he", "thong", "mau", "kiem", "tra",
}

ABSENT_TOPICS = [
    "blockchain", "mat khau admin", "CCCD cua sinh vien", "benh vien",
    "gia co phieu Apple", "dia chi nha rieng", "so tai khoan ngan hang",
    "ma OTP", "du lieu y te benh nhan", "hop dong lao dong",
    "bitcoin", "ma nguon bi mat", "lich bay quoc te", "diem thi IELTS",
    "so dien thoai ca nhan", "the tin dung", "API key", "du lieu GPS",
    "luong nhan vien", "thong tin bao hiem",
]


def fold(value: object) -> str:
    text = str(value or "").replace("Đ", "D").replace("đ", "d")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.lower()
    text = re.sub(r"[^a-z0-9\u3040-\u30ff\u3400-\u9fff]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def compact(value: object, limit: int = 900) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text[:limit]


def is_fallback(answer: str) -> bool:
    folded = fold(answer)
    return any(
        phrase in folded
        for phrase in [
            "khong tim thay", "khong co thong tin", "khong du thong tin",
            "khong the trich", "khong nam trong tai lieu", "khong de cap",
            "insufficient", "not found", "outside context",
        ]
    )


def is_negative_answer(answer: str) -> bool:
    folded = fold(answer)
    return is_fallback(answer) or bool(
        re.search(r"\b(khong|chua|khong co|khong tim thay|khong de cap)\b", folded)
    )


def has_term(text: str, term: str) -> bool:
    folded_text = fold(text)
    folded_term = fold(term)
    return bool(folded_term and folded_term in folded_text)


def count_term_hits(text: str, terms: list[str]) -> int:
    return sum(1 for term in terms if has_term(text, term))


def source_matches(sources: list[str], expected_files: list[str]) -> bool:
    if not expected_files:
        return True
    folded_sources = fold("\n".join(str(source) for source in sources))
    return all(fold(name) in folded_sources for name in expected_files)


def source_matches_any(sources: list[str], expected_files: list[str]) -> bool:
    if not expected_files:
        return True
    folded_sources = fold("\n".join(str(source) for source in sources))
    return any(fold(name) in folded_sources for name in expected_files)


def useful_lines(text: str, *, max_lines: int = 12) -> list[str]:
    lines: list[str] = []
    for raw in str(text or "").splitlines():
        line = re.sub(r"\s+", " ", raw).strip(" -•\t")
        if len(line) < 8:
            continue
        folded = fold(line)
        if not folded:
            continue
        if folded.startswith(("file ", "layout ", "reading order", "slide blocks")):
            continue
        if folded in {"overview", "tong quan"}:
            continue
        if line not in lines:
            lines.append(line[:260])
        if len(lines) >= max_lines:
            break
    return lines


def extract_keywords(text: str, *, limit: int = 8) -> list[str]:
    folded_text = fold(text)
    raw_tokens = re.findall(r"[a-z0-9\u3040-\u30ff\u3400-\u9fff]{2,}", folded_text)
    tokens: list[str] = []
    for token in raw_tokens:
        if token in STOPWORDS:
            continue
        if token.isdigit() and len(token) < 2:
            continue
        if len(token) < 3 and not re.search(r"\d", token):
            continue
        tokens.append(token)
    counts = Counter(tokens)
    keywords = [token for token, _ in counts.most_common(limit * 3)]

    numbers = re.findall(r"\b\d{2,}(?:[.,:]\d+)?\b", str(text or ""))
    merged: list[str] = []
    for item in [*numbers[:3], *keywords]:
        if item not in merged and item not in STOPWORDS:
            merged.append(item)
        if len(merged) >= limit:
            break
    return merged


def extension_type(name: str) -> str:
    suffix = Path(name).suffix.lower().lstrip(".")
    if suffix == "pdf":
        return "PDF"
    if suffix in {"jpg", "jpeg", "png"}:
        return suffix.upper()
    return suffix.upper() or "UNKNOWN"


@dataclass(slots=True)
class FileProfile:
    document_id: str
    original_name: str
    stored_path: str
    file_type: str
    chunks: list[Document]
    chunk_count: int
    text: str
    title: str
    terms: list[str]
    lines: list[str]


def doc_sort_key(profile: FileProfile) -> tuple[str, str]:
    return (profile.file_type, profile.original_name.lower())


def build_profiles(container: Any, username: str, chat_id: str) -> list[FileProfile]:
    documents = container.workspace_service.list_documents(username=username, chat_id=chat_id)
    profiles: list[FileProfile] = []
    for stored in documents:
        chunks = container.vector_store_repository.list_documents(
            {"owner": username, "chat_id": chat_id, "source": stored.stored_path},
            limit=None,
        )
        text = "\n".join(doc.page_content for doc in chunks if str(doc.page_content or "").strip())
        metadata_lines: list[str] = []
        for doc in chunks[:10]:
            for key in ("title", "section_title", "structure_path", "citation_hint", "sheet_name"):
                value = str(doc.metadata.get(key) or "").strip()
                if value and value not in metadata_lines:
                    metadata_lines.append(value)
        lines = useful_lines("\n".join(metadata_lines + useful_lines(text, max_lines=20)), max_lines=14)
        title = lines[0] if lines else Path(stored.original_name).stem
        terms = extract_keywords("\n".join(lines[:8]) or text, limit=8)
        profiles.append(
            FileProfile(
                document_id=stored.document_id,
                original_name=stored.original_name,
                stored_path=stored.stored_path,
                file_type=extension_type(stored.original_name),
                chunks=chunks,
                chunk_count=len(chunks),
                text=text,
                title=title,
                terms=terms,
                lines=lines,
            )
        )
    return profiles


def make_case(
    case_id: str,
    group: str,
    profiles: list[FileProfile],
    question: str,
    expected_answer: str,
    expected_terms: list[str],
    *,
    kind: str = "positive",
    forbidden_terms: list[str] | None = None,
) -> dict[str, Any]:
    return {
        "id": case_id,
        "group": group,
        "files": [p.original_name for p in profiles],
        "file_type": "+".join(sorted({p.file_type for p in profiles})),
        "selected_document_ids": [p.document_id for p in profiles],
        "question": question,
        "expected_answer": expected_answer,
        "expected_terms": expected_terms[:10],
        "forbidden_terms": forbidden_terms or [],
        "kind": kind,
    }


def build_direct_cases(profiles: list[FileProfile], target: int) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    idx = 1
    for profile in profiles:
        terms = profile.terms[:5] or [Path(profile.original_name).stem]
        expected = profile.title or "; ".join(terms)
        cases.append(make_case(
            f"DIR-{idx:03d}", "Direct", [profile],
            f'Tài liệu "{profile.original_name}" nói về nội dung gì?',
            expected, terms,
        ))
        idx += 1
        cases.append(make_case(
            f"DIR-{idx:03d}", "Direct", [profile],
            f'Nêu một thông tin cụ thể xuất hiện trong file "{profile.original_name}".',
            expected, terms,
        ))
        idx += 1
    pointer = 0
    rich = sorted(profiles, key=lambda p: p.chunk_count, reverse=True)
    while len(cases) < target:
        profile = rich[pointer % len(rich)]
        line = profile.lines[pointer % max(1, len(profile.lines))] if profile.lines else profile.title
        terms = extract_keywords(line, limit=5) or profile.terms[:5]
        cases.append(make_case(
            f"DIR-{idx:03d}", "Direct", [profile],
            f'Trong file "{profile.original_name}", ý hoặc mục nào liên quan đến "{terms[0] if terms else Path(profile.original_name).stem}"?',
            line, terms,
        ))
        idx += 1
        pointer += 1
    return cases[:target]


def build_summary_cases(profiles: list[FileProfile], target: int) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    idx = 1
    for profile in profiles:
        terms = profile.terms[:6] or extract_keywords(profile.text, limit=6)
        expected = f"Tóm tắt có các ý chính: {', '.join(terms[:4])}"
        cases.append(make_case(
            f"SUM-{idx:03d}", "Summary", [profile],
            f'Tóm tắt ngắn file "{profile.original_name}" trong 3 ý.',
            expected, terms,
        ))
        idx += 1
    rich = sorted(profiles, key=lambda p: p.chunk_count, reverse=True)
    pointer = 0
    while len(cases) < target:
        profile = rich[pointer % len(rich)]
        terms = profile.terms[:8]
        cases.append(make_case(
            f"SUM-{idx:03d}", "Summary", [profile],
            f'Tổng hợp các nội dung chính và điểm đáng chú ý trong "{profile.original_name}".',
            f"Tổng hợp đúng các ý liên quan: {', '.join(terms[:5])}",
            terms,
        ))
        idx += 1
        pointer += 1
    return cases[:target]


def build_comparison_cases(profiles: list[FileProfile], target: int) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    sorted_profiles = sorted(profiles, key=doc_sort_key)
    pairs: list[tuple[FileProfile, FileProfile]] = []
    for i, left in enumerate(sorted_profiles):
        right = sorted_profiles[(i + 7) % len(sorted_profiles)]
        if left.document_id != right.document_id:
            pairs.append((left, right))
    for left in sorted_profiles:
        for right in sorted_profiles:
            if left.file_type != right.file_type and left.document_id != right.document_id:
                pairs.append((left, right))
            if len(pairs) >= target * 2:
                break
        if len(pairs) >= target * 2:
            break

    seen: set[tuple[str, str]] = set()
    idx = 1
    for left, right in pairs:
        key = tuple(sorted([left.document_id, right.document_id]))
        if key in seen:
            continue
        seen.add(key)
        terms = (left.terms[:3] + right.terms[:3])[:8]
        cases.append(make_case(
            f"CMP-{idx:03d}", "Comparison", [left, right],
            f'So sánh nội dung của "{left.original_name}" và "{right.original_name}".',
            f"Cần nêu được nội dung của cả hai file: {left.title}; {right.title}",
            terms,
        ))
        idx += 1
        if len(cases) >= target:
            break
    return cases


def structured_rows_from_profile(profile: FileProfile) -> list[tuple[str, int, dict[str, Any]]]:
    rows: list[tuple[str, int, dict[str, Any]]] = []
    for doc in profile.chunks:
        sheet = str(doc.metadata.get("sheet_name") or doc.metadata.get("sheet") or "").strip()
        raw_rows = doc.metadata.get("structured_rows")
        if isinstance(raw_rows, list):
            for row in raw_rows:
                if not isinstance(row, dict):
                    continue
                values = row.get("values")
                if isinstance(values, dict) and values:
                    try:
                        row_number = int(row.get("row_number") or row.get("index") or 0)
                    except (TypeError, ValueError):
                        row_number = 0
                    rows.append((sheet, row_number, values))
    return rows


def build_table_cases(profiles: list[FileProfile], target: int) -> list[dict[str, Any]]:
    spreadsheets = [p for p in profiles if p.file_type in {"XLSX", "XLS"}]
    cases: list[dict[str, Any]] = []
    idx = 1
    for profile in spreadsheets:
        sheets = []
        for doc in profile.chunks:
            sheet = str(doc.metadata.get("sheet_name") or "").strip()
            if sheet and sheet not in sheets:
                sheets.append(sheet)
        if sheets:
            cases.append(make_case(
                f"TAB-{idx:03d}", "Table/Numeric", [profile],
                f'Workbook "{profile.original_name}" có bao nhiêu sheet và tên gì?',
                f"{len(sheets)} sheet: {', '.join(sheets)}",
                [str(len(sheets)), *sheets[:5]],
            ))
            idx += 1
        rows = structured_rows_from_profile(profile)
        for sheet, row_number, values in rows[:12]:
            items = [(str(k), str(v)) for k, v in values.items() if str(v).strip()]
            if not items:
                continue
            terms = [value for _, value in items[:4]]
            expected = "; ".join(f"{k}: {v}" for k, v in items[:5])
            row_hint = f"dòng {row_number}" if row_number else "một dòng dữ liệu"
            cases.append(make_case(
                f"TAB-{idx:03d}", "Table/Numeric", [profile],
                f'Trong file "{profile.original_name}", ở sheet "{sheet}", {row_hint} có thông tin gì?',
                expected,
                terms,
            ))
            idx += 1
            if len(cases) >= target:
                return cases[:target]
    pointer = 0
    while len(cases) < target and spreadsheets:
        profile = spreadsheets[pointer % len(spreadsheets)]
        terms = profile.terms[:6]
        cases.append(make_case(
            f"TAB-{idx:03d}", "Table/Numeric", [profile],
            f'Liệt kê các cột hoặc số liệu chính trong file bảng "{profile.original_name}".',
            f"Các cột/số liệu chính liên quan: {', '.join(terms)}",
            terms,
        ))
        idx += 1
        pointer += 1
    return cases[:target]


def build_slide_cases(profiles: list[FileProfile], target: int) -> list[dict[str, Any]]:
    decks = [p for p in profiles if p.file_type == "PPTX"]
    cases: list[dict[str, Any]] = []
    idx = 1
    for profile in decks:
        slide_docs = []
        seen: set[int] = set()
        for doc in profile.chunks:
            raw = doc.metadata.get("slide_number") or doc.metadata.get("slide")
            try:
                slide_no = int(raw)
            except (TypeError, ValueError):
                continue
            if slide_no in seen:
                continue
            seen.add(slide_no)
            slide_docs.append((slide_no, doc))
        for slide_no, doc in slide_docs[:8]:
            terms = extract_keywords(doc.page_content, limit=6) or profile.terms[:6]
            expected = "; ".join(useful_lines(doc.page_content, max_lines=3)) or profile.title
            cases.append(make_case(
                f"SLD-{idx:03d}", "Slide/Presentation", [profile],
                f'Slide {slide_no} của file "{profile.original_name}" nói về nội dung gì?',
                expected,
                terms,
            ))
            idx += 1
            if len(cases) >= target:
                return cases[:target]
    return cases[:target]


def build_ocr_cases(profiles: list[FileProfile], target: int) -> list[dict[str, Any]]:
    ocr_profiles = [
        p for p in profiles
        if p.file_type in {"PNG", "JPG", "JPEG"} or "scan" in p.original_name.lower()
    ]
    cases: list[dict[str, Any]] = []
    idx = 1
    for profile in ocr_profiles:
        terms = profile.terms[:6] or extract_keywords(profile.text, limit=6)
        cases.append(make_case(
            f"OCR-{idx:03d}", "OCR/Vision", [profile],
            f'Trong ảnh hoặc file scan "{profile.original_name}", nội dung chữ chính là gì?',
            profile.title or f"Nội dung OCR/Vision chứa: {', '.join(terms[:4])}",
            terms,
        ))
        idx += 1
        if len(cases) >= target:
            return cases
    pointer = 0
    while len(cases) < target and ocr_profiles:
        profile = ocr_profiles[pointer % len(ocr_profiles)]
        terms = profile.terms[:6]
        cases.append(make_case(
            f"OCR-{idx:03d}", "OCR/Vision", [profile],
            f'Hãy trích xuất các nhãn hoặc ý chính nhìn thấy trong "{profile.original_name}".',
            profile.title,
            terms,
        ))
        idx += 1
        pointer += 1
    return cases[:target]


def build_negative_cases(profiles: list[FileProfile], target: int) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    ordered = sorted(profiles, key=lambda p: p.original_name.lower())
    for idx in range(target):
        profile = ordered[idx % len(ordered)]
        topic = ABSENT_TOPICS[idx % len(ABSENT_TOPICS)]
        cases.append(make_case(
            f"NEG-{idx + 1:03d}", "Negative/Hallucination", [profile],
            f'Trong file "{profile.original_name}" có thông tin về {topic} không?',
            "Không tìm thấy thông tin trong tài liệu.",
            ["không", "không tìm thấy", "không có thông tin"],
            kind="negative",
            forbidden_terms=[topic],
        ))
    return cases


def build_cases(profiles: list[FileProfile]) -> list[dict[str, Any]]:
    cases = []
    cases.extend(build_direct_cases(profiles, 90))
    cases.extend(build_summary_cases(profiles, 60))
    cases.extend(build_comparison_cases(profiles, 40))
    cases.extend(build_table_cases(profiles, 40))
    cases.extend(build_slide_cases(profiles, 30))
    cases.extend(build_ocr_cases(profiles, 20))
    cases.extend(build_negative_cases(profiles, 20))
    if len(cases) != 300:
        raise RuntimeError(f"Expected 300 cases, got {len(cases)}")
    return cases


def ask_case(container: Any, username: str, chat_id: str, case: dict[str, Any]) -> tuple[float, str, list[str]]:
    question = str(case["question"])
    selected_ids = list(case["selected_document_ids"])
    routing = _resolve_ask_routing(
        username=username,
        chat_id=chat_id,
        question=question,
        selected_document_ids=selected_ids,
        workspace_service=container.workspace_service,
    )
    if routing.clarification_answer:
        return 0.0, routing.clarification_answer, []

    _ensure_workspace_scope_indexed(
        username=username,
        chat_id=chat_id,
        metadata_filter=routing.metadata_filter,
        workspace_service=container.workspace_service,
        vector_store_repository=container.vector_store_repository,
        ingestion_service=container.ingestion_service,
    )

    effective_question = routing.effective_question
    if routing.prefer_combined_answer and routing.scoped_documents and len(routing.scoped_documents) > 1:
        effective_question = _inject_document_mapping_into_question(
            routing.effective_question,
            routing.scoped_documents,
            routing.scoped_document_numbers,
        )

    def _timeout_handler(_signum: int, _frame: Any) -> None:
        raise TimeoutError("ask_case_timeout")

    started = time.perf_counter()
    old_handler = signal.signal(signal.SIGALRM, _timeout_handler)
    signal.alarm(75)
    try:
        result = container.question_answering_service.ask(
            effective_question,
            metadata_filter=routing.metadata_filter,
            top_k=8,
        )
        latency = time.perf_counter() - started
        return latency, result.answer, result.sources
    except TimeoutError:
        latency = time.perf_counter() - started
        return latency, "TIMEOUT: hệ thống không trả lời trong 75 giây.", []
    except Exception as exc:
        latency = time.perf_counter() - started
        return latency, f"ERROR: {type(exc).__name__}: {exc}", []
    finally:
        signal.alarm(0)
        signal.signal(signal.SIGALRM, old_handler)


def evaluate_case(case: dict[str, Any], answer: str, sources: list[str]) -> dict[str, Any]:
    terms = list(case.get("expected_terms") or [])
    forbidden = list(case.get("forbidden_terms") or [])
    expected_files = list(case.get("files") or [])
    kind = str(case.get("kind") or "positive")

    term_hits = count_term_hits(answer, terms)
    term_total = max(1, min(len(terms), 5))
    citation_ok = source_matches(sources, expected_files)
    citation_partial = source_matches_any(sources, expected_files)
    retrieval_ok = citation_ok or (kind == "negative" and is_negative_answer(answer))
    fallback = is_fallback(answer)
    contains_forbidden = any(has_term(answer, item) for item in forbidden if item)
    visual_noise = bool(re.search(r"```mermaid|flowchart|mindmap", answer, re.IGNORECASE))

    if kind == "negative":
        answer_ok = is_negative_answer(answer)
        hallucination = not answer_ok
        if answer_ok and (citation_ok or not sources):
            score = 1.0
            answer_status = "Đúng"
        elif answer_ok:
            score = 0.5
            answer_status = "Một phần"
        else:
            score = 0.0
            answer_status = "Sai"
    else:
        if fallback:
            answer_ok = False
            answer_status = "Sai"
            score = 0.0
        elif term_hits >= min(2, term_total) and citation_ok and not contains_forbidden:
            answer_ok = True
            answer_status = "Đúng"
            score = 1.0
        elif (term_hits >= 1 or citation_partial) and not contains_forbidden:
            answer_ok = False
            answer_status = "Một phần"
            score = 0.5
        else:
            answer_ok = False
            answer_status = "Sai"
            score = 0.0
        hallucination = bool(contains_forbidden or (score == 0 and not fallback and not citation_partial))

    if citation_ok:
        citation_status = "Đúng"
    elif citation_partial:
        citation_status = "Một phần"
    elif kind == "negative" and not sources and is_negative_answer(answer):
        citation_status = "Đúng"
    else:
        citation_status = "Sai"

    retrieval_status = "Đúng" if retrieval_ok else ("Một phần" if citation_partial else "Sai")

    if hallucination:
        severity = "Critical"
    elif score == 0:
        severity = "High"
    elif score == 0.5 or citation_status != "Đúng":
        severity = "Medium"
    elif visual_noise:
        severity = "Low"
    else:
        severity = "None"

    notes: list[str] = []
    if fallback and kind != "negative":
        notes.append("Fallback dù case có đáp án kỳ vọng.")
    if citation_status != "Đúng":
        notes.append("Citation không bao phủ đúng file kỳ vọng.")
    if retrieval_status != "Đúng":
        notes.append("Retrieval/citation có dấu hiệu sai hoặc thiếu nguồn.")
    if hallucination:
        notes.append("Có dấu hiệu suy đoán hoặc trả lời ngoài tài liệu.")
    if visual_noise:
        notes.append("Có sinh sơ đồ/mermaid khi câu hỏi không yêu cầu rõ.")
    if not notes:
        notes.append("Đạt kỳ vọng theo bộ từ khóa chuẩn.")

    fix = "Không cần fix."
    if severity in {"Critical", "High"}:
        if "OCR" in str(case["group"]) or case["file_type"] in {"PNG", "JPG", "JPEG"}:
            fix = "Rà soát OCR/Vision extraction và metadata ảnh; tăng kiểm tra fallback cho câu hỏi ảnh."
        elif "Table" in str(case["group"]) or "XLS" in str(case["file_type"]):
            fix = "Tăng structured extraction cho Excel/bảng, giữ header/sheet/row metadata và ưu tiên table retriever."
        elif "Slide" in str(case["group"]) or "PPTX" in str(case["file_type"]):
            fix = "Bổ sung metadata slide/title/shape, ưu tiên slide filter và citation theo slide."
        elif "Comparison" in str(case["group"]):
            fix = "Bảo toàn context từ từng tài liệu được chọn trước reranking, tránh lệch về một nguồn."
        elif kind == "negative":
            fix = "Siết missing-evidence gate và prompt fallback cho câu hỏi không có thông tin."
        else:
            fix = "Kiểm tra chunking/retrieval/reranking và ngưỡng fallback cho file liên quan."
    elif severity == "Medium":
        fix = "Tinh chỉnh citation selection và reranking để nguồn trả về sát hơn."

    return {
        "actual_answer": compact(answer, 1200),
        "sources": " | ".join(sources),
        "retrieval": retrieval_status,
        "citation": citation_status,
        "answer_status": answer_status,
        "score": score,
        "hallucination": "Có" if hallucination else "Không",
        "severity": severity,
        "notes": " ".join(notes),
        "fix": fix,
        "term_hits": term_hits,
        "latency_ms": None,
    }


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def write_xlsx(path: Path, detail_rows: list[dict[str, Any]], file_rows: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    def clean_cell(value: Any) -> Any:
        if not isinstance(value, str):
            return value
        return re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", " ", value)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    for key, value in summary.items():
        ws.append([clean_cell(key), clean_cell(value)])
    ws["A1"].font = ws["B1"].font = Font(bold=True)

    ws2 = wb.create_sheet("300 Test Cases")
    headers = [
        "ID", "File", "Loại file", "Nhóm", "Câu hỏi", "Đáp án chuẩn",
        "Câu trả lời hệ thống", "Nguồn/Citation", "Retrieval", "Citation",
        "Answer", "Điểm", "Hallucination", "Mức độ lỗi", "Ghi chú", "Đề xuất fix",
        "Latency ms",
    ]
    ws2.append(headers)
    for row in detail_rows:
        ws2.append([
            clean_cell(row["ID"]), clean_cell(row["File"]), clean_cell(row["Loại file"]),
            clean_cell(row["Nhóm"]), clean_cell(row["Câu hỏi"]),
            clean_cell(row["Đáp án chuẩn"]), clean_cell(row["Câu trả lời hệ thống"]),
            clean_cell(row["Nguồn/Citation"]),
            clean_cell(row["Retrieval"]), clean_cell(row["Citation"]), clean_cell(row["Answer"]), row["Điểm"],
            clean_cell(row["Hallucination"]), clean_cell(row["Mức độ lỗi"]), clean_cell(row["Ghi chú"]),
            clean_cell(row["Đề xuất fix"]),
            row["Latency ms"],
        ])

    ws3 = wb.create_sheet("Files")
    ws3.append(["File", "Loại file", "Chunks", "Ingest", "Ghi chú"])
    for row in file_rows:
        ws3.append([clean_cell(row["file"]), clean_cell(row["type"]), row["chunks"], clean_cell(row["ingest"]), clean_cell(row["notes"])])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F766E")
        for column_cells in sheet.columns:
            width = min(60, max(12, max(len(str(cell.value or "")) for cell in column_cells[:60]) + 2))
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def percent(value: float) -> str:
    return f"{value * 100:.2f}%"


def build_markdown(
    path: Path,
    summary: dict[str, Any],
    detail_rows: list[dict[str, Any]],
    file_rows: list[dict[str, Any]],
    format_stats: dict[str, dict[str, Any]],
    top_errors: list[dict[str, Any]],
) -> None:
    lines: list[str] = []
    lines.append("# Manual Test Report - AI Document Chat/RAG\n")
    lines.append("## A. Tổng quan kiểm thử\n")
    for key, value in summary.items():
        lines.append(f"- **{key}:** {value}")
    lines.append("\n### File ingestion\n")
    lines.append("| File | Loại | Chunks | Ingest | Ghi chú |")
    lines.append("|---|---:|---:|---|---|")
    for row in file_rows:
        lines.append(f"| {row['file']} | {row['type']} | {row['chunks']} | {row['ingest']} | {row['notes']} |")

    lines.append("\n## B. Bảng kết quả chi tiết 300 câu hỏi\n")
    lines.append("Bảng đầy đủ nằm trong `manual_test_report.xlsx`. Dưới đây là 40 dòng đầu và toàn bộ lỗi nằm trong `failed_cases.csv`.\n")
    lines.append("| ID | File | Nhóm | Retrieval | Citation | Điểm | Hallucination | Mức độ lỗi | Ghi chú |")
    lines.append("|---|---|---|---|---|---:|---|---|---|")
    for row in detail_rows[:40]:
        lines.append(
            f"| {row['ID']} | {row['File']} | {row['Nhóm']} | {row['Retrieval']} | "
            f"{row['Citation']} | {row['Điểm']} | {row['Hallucination']} | {row['Mức độ lỗi']} | {row['Ghi chú'][:120]} |"
        )

    lines.append("\n## C. Phân tích lỗi theo loại file\n")
    lines.append("| Loại file | Số case | Điểm TB | Full correct | Partial | Fail | Hallucination | Nhận xét |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|---|")
    for ftype, stats in sorted(format_stats.items()):
        lines.append(
            f"| {ftype} | {stats['cases']} | {stats['avg_score']:.2f} | {stats['full']} | "
            f"{stats['partial']} | {stats['fail']} | {stats['hallucination']} | {stats['comment']} |"
        )

    lines.append("\n## D. Top lỗi nghiêm trọng nhất\n")
    if not top_errors:
        lines.append("Không ghi nhận lỗi nghiêm trọng trong bộ test tự động.")
    else:
        for idx, row in enumerate(top_errors[:10], start=1):
            lines.append(
                f"{idx}. **{row['Mức độ lỗi']} - {row['ID']} ({row['File']})**\n"
                f"   - Câu hỏi: {row['Câu hỏi']}\n"
                f"   - Trả lời hệ thống: {row['Câu trả lời hệ thống'][:300]}\n"
                f"   - Nguyên nhân/ghi chú: {row['Ghi chú']}\n"
                f"   - Đề xuất: {row['Đề xuất fix']}\n"
            )

    lines.append("\n## E. Đề xuất cải thiện hệ thống\n")
    lines.extend([
        "- **File parser:** thêm regression fixture cho từng định dạng và so sánh số chunk/metadata sau mỗi lần sửa.",
        "- **OCR/Vision:** tăng kiểm thử ảnh nhiễu, ảnh nhiều chữ, PDF scan; lưu OCR confidence nếu provider hỗ trợ.",
        "- **Excel extraction:** giữ đầy đủ sheet/header/row/range/formula cached value; ưu tiên structured table answer cho câu hỏi số liệu.",
        "- **PPTX extraction:** tăng metadata slide/title/shape/table/image; citation theo slide phải ưu tiên đúng slide được hỏi.",
        "- **Chunking:** kiểm soát chunk không cắt mất heading/bảng; bổ sung parent metadata cho các chunk con.",
        "- **Metadata:** chuẩn hóa source, page, slide, sheet, row_index, content_type để citation ổn định.",
        "- **Embedding:** tiếp tục dùng MiniLM-L12-v2 cho độ trễ thấp; cân nhắc A/B với model mạnh hơn cho OCR và tiếng Nhật.",
        "- **Retrieval/reranking:** bảo toàn context theo từng tài liệu khi multi-select; thêm reranker chuyên cho bảng/slide.",
        "- **Citation:** giới hạn nguồn theo câu hỏi cụ thể; loại citation trùng và nguồn nhiễu.",
        "- **Prompt/fallback:** giữ nguyên tắc trả lời ngắn, không bịa; negative question phải trả lời không tìm thấy khi thiếu bằng chứng.",
    ])
    path.write_text("\n".join(lines), encoding="utf-8")


def build_improvement_plan(path: Path, top_errors: list[dict[str, Any]], format_stats: dict[str, dict[str, Any]]) -> None:
    lines = [
        "# Improvement Plan",
        "",
        "## P0 - Critical / High",
    ]
    critical_high = [row for row in top_errors if row["Mức độ lỗi"] in {"Critical", "High"}]
    if not critical_high:
        lines.append("- Chưa có lỗi P0 rõ ràng trong bộ 300 case tự động; tiếp tục mở rộng câu hỏi thủ công khó hơn.")
    else:
        for row in critical_high[:12]:
            lines.append(f"- **{row['ID']} - {row['File']}**: {row['Ghi chú']} Fix: {row['Đề xuất fix']}")

    lines.extend([
        "",
        "## P1 - Retrieval / Citation",
        "- Thêm unit test cho citation theo page/slide/sheet/row khi câu hỏi có location hint.",
        "- Thêm scoring bắt buộc mỗi selected document có ít nhất một context khi hỏi so sánh.",
        "- Gắn metadata `content_type` nhất quán cho OCR, slide image, spreadsheet row/table.",
        "",
        "## P2 - Format Robustness",
    ])
    for ftype, stats in sorted(format_stats.items(), key=lambda item: item[1]["avg_score"]):
        lines.append(f"- **{ftype}**: avg_score={stats['avg_score']:.2f}, fail={stats['fail']}, hallucination={stats['hallucination']}. {stats['comment']}")

    lines.extend([
        "",
        "## P3 - Continuous Evaluation",
        "- Lưu bộ 300 case này làm regression dataset.",
        "- Chạy lại sau mỗi thay đổi parser/chunking/retrieval.",
        "- Tách thêm benchmark Hit@1/Hit@5/MRR từ expected source metadata.",
    ])
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--username", required=True)
    parser.add_argument("--chat-id", required=True)
    parser.add_argument("--out-dir", default="/app/tmp/rag_audit_300")
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print("audit_start", flush=True)
    container = build_container(get_settings())
    print("container_ready", flush=True)
    profiles = build_profiles(container, args.username, args.chat_id)
    print(f"profiles={len(profiles)}", flush=True)
    cases = build_cases(profiles)
    print(f"cases={len(cases)}", flush=True)

    profile_by_id = {profile.document_id: profile for profile in profiles}
    detail_rows: list[dict[str, Any]] = []
    raw_results: list[dict[str, Any]] = []
    checkpoint_path = out_dir / "manual_test_results_checkpoint.json"
    if checkpoint_path.exists():
        checkpoint = json.loads(checkpoint_path.read_text(encoding="utf-8"))
        if int(checkpoint.get("completed") or 0) == 300:
            raw_results = list(checkpoint.get("results") or [])
            rebuilt_rows: list[dict[str, Any]] = []
            for item in raw_results:
                if "case" not in item or "row" not in item:
                    continue
                case = item["case"]
                old_row = item["row"]
                answer = str(old_row.get("Câu trả lời hệ thống") or "")
                sources = str(old_row.get("Nguồn/Citation") or "").split(" | ") if old_row.get("Nguồn/Citation") else []
                evaluation = evaluate_case(case, answer, sources)
                evaluation["latency_ms"] = old_row.get("Latency ms")
                rebuilt_rows.append({
                    "ID": case["id"],
                    "File": ", ".join(case["files"]),
                    "Loại file": case["file_type"],
                    "Nhóm": case["group"],
                    "Câu hỏi": case["question"],
                    "Đáp án chuẩn": case["expected_answer"],
                    "Câu trả lời hệ thống": evaluation["actual_answer"],
                    "Nguồn/Citation": evaluation["sources"],
                    "Retrieval": evaluation["retrieval"],
                    "Citation": evaluation["citation"],
                    "Answer": evaluation["answer_status"],
                    "Điểm": evaluation["score"],
                    "Hallucination": evaluation["hallucination"],
                    "Mức độ lỗi": evaluation["severity"],
                    "Ghi chú": evaluation["notes"],
                    "Đề xuất fix": evaluation["fix"],
                    "Latency ms": evaluation["latency_ms"],
                })
                item["evaluation"] = evaluation
                item["row"] = rebuilt_rows[-1]
            detail_rows = rebuilt_rows
            print("loaded_checkpoint=300", flush=True)

    if not detail_rows:
        for index, case in enumerate(cases, start=1):
            latency, answer, sources = ask_case(container, args.username, args.chat_id, case)
            evaluation = evaluate_case(case, answer, sources)
            evaluation["latency_ms"] = round(latency * 1000, 2)
            row = {
                "ID": case["id"],
                "File": ", ".join(case["files"]),
                "Loại file": case["file_type"],
                "Nhóm": case["group"],
                "Câu hỏi": case["question"],
                "Đáp án chuẩn": case["expected_answer"],
                "Câu trả lời hệ thống": evaluation["actual_answer"],
                "Nguồn/Citation": evaluation["sources"],
                "Retrieval": evaluation["retrieval"],
                "Citation": evaluation["citation"],
                "Answer": evaluation["answer_status"],
                "Điểm": evaluation["score"],
                "Hallucination": evaluation["hallucination"],
                "Mức độ lỗi": evaluation["severity"],
                "Ghi chú": evaluation["notes"],
                "Đề xuất fix": evaluation["fix"],
                "Latency ms": evaluation["latency_ms"],
            }
            detail_rows.append(row)
            raw_results.append({"case": case, "evaluation": evaluation, "row": row})
            if index % 10 == 0:
                checkpoint_path.write_text(
                    json.dumps({"completed": index, "results": raw_results}, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
            print(f"{index:03d}/300 {case['id']} score={evaluation['score']} severity={evaluation['severity']} latency_ms={evaluation['latency_ms']}", flush=True)

    total = len(detail_rows)
    full = sum(1 for row in detail_rows if float(row["Điểm"]) == 1.0)
    partial = sum(1 for row in detail_rows if float(row["Điểm"]) == 0.5)
    fail = sum(1 for row in detail_rows if float(row["Điểm"]) == 0.0)
    retrieval_full = sum(1 for row in detail_rows if row["Retrieval"] == "Đúng")
    citation_full = sum(1 for row in detail_rows if row["Citation"] == "Đúng")
    hallucinations = sum(1 for row in detail_rows if row["Hallucination"] == "Có")
    negative_rows = [row for row in detail_rows if row["Nhóm"] == "Negative/Hallucination"]
    no_answer_correct = sum(1 for row in negative_rows if float(row["Điểm"]) >= 0.5)
    wrong_or_no_answer_denominator = len(negative_rows) + fail
    hallucination_rate = hallucinations / max(1, wrong_or_no_answer_denominator)
    latencies = [float(row["Latency ms"]) for row in detail_rows if row["Latency ms"] not in {"", None}]

    file_rows: list[dict[str, Any]] = []
    for profile in profiles:
        ingest_ok = profile.chunk_count > 0
        file_rows.append({
            "file": profile.original_name,
            "type": profile.file_type,
            "chunks": profile.chunk_count,
            "ingest": "Pass" if ingest_ok else "Fail",
            "notes": "Có chunk và metadata truy xuất." if ingest_ok else "Không có chunk sau ingest.",
        })

    format_stats: dict[str, dict[str, Any]] = {}
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in detail_rows:
        for ftype in str(row["Loại file"]).split("+"):
            grouped[ftype].append(row)
    for ftype, rows in grouped.items():
        avg_score = statistics.mean(float(row["Điểm"]) for row in rows) if rows else 0.0
        f_full = sum(1 for row in rows if float(row["Điểm"]) == 1.0)
        f_partial = sum(1 for row in rows if float(row["Điểm"]) == 0.5)
        f_fail = sum(1 for row in rows if float(row["Điểm"]) == 0.0)
        f_hall = sum(1 for row in rows if row["Hallucination"] == "Có")
        if avg_score >= 0.9:
            comment = "Hoạt động tốt trong bộ test."
        elif avg_score >= 0.75:
            comment = "Tương đối ổn, cần rà soát các case sai/citation."
        else:
            comment = "Còn yếu, cần ưu tiên cải thiện parser/retrieval."
        format_stats[ftype] = {
            "cases": len(rows),
            "avg_score": avg_score,
            "full": f_full,
            "partial": f_partial,
            "fail": f_fail,
            "hallucination": f_hall,
            "comment": comment,
        }

    ingestion_success = sum(1 for row in file_rows if row["ingest"] == "Pass") / max(1, len(file_rows))
    summary = OrderedDict([
        ("Tổng số file đã test", len(file_rows)),
        ("Tổng số câu hỏi", total),
        ("Các loại file đã test", ", ".join(sorted({row["type"] for row in file_rows}))),
        ("File Ingestion Success Rate", percent(ingestion_success)),
        ("Answer Accuracy", f"{full}/{total} = {percent(full / total)}"),
        ("Partial Accuracy", f"{partial}/{total} = {percent(partial / total)}"),
        ("Fail Count", fail),
        ("Retrieval Correctness", f"{retrieval_full}/{total} = {percent(retrieval_full / total)}"),
        ("Citation Correctness", f"{citation_full}/{total} = {percent(citation_full / total)}"),
        ("Hallucination Rate", f"{hallucinations}/{wrong_or_no_answer_denominator} = {percent(hallucination_rate)}"),
        ("No Answer Correctness", f"{no_answer_correct}/20 = {percent(no_answer_correct / 20)}"),
        ("Mean Latency", f"{statistics.mean(latencies):.2f} ms" if latencies else "N/A"),
    ])

    failed_rows = [
        row for row in detail_rows
        if float(row["Điểm"]) < 1.0 or row["Retrieval"] != "Đúng" or row["Citation"] != "Đúng" or row["Hallucination"] == "Có"
    ]
    severity_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3, "None": 4}
    top_errors = sorted(
        [row for row in failed_rows if row["Mức độ lỗi"] != "None"],
        key=lambda row: (severity_order.get(row["Mức độ lỗi"], 9), float(row["Điểm"]), -float(row["Latency ms"] or 0)),
    )

    detail_fieldnames = [
        "ID", "File", "Loại file", "Nhóm", "Câu hỏi", "Đáp án chuẩn",
        "Câu trả lời hệ thống", "Nguồn/Citation", "Retrieval", "Citation",
        "Answer", "Điểm", "Hallucination", "Mức độ lỗi", "Ghi chú", "Đề xuất fix", "Latency ms",
    ]
    write_csv(out_dir / "manual_test_cases_300.csv", detail_rows, detail_fieldnames)
    write_csv(out_dir / "failed_cases.csv", failed_rows, detail_fieldnames)
    write_xlsx(out_dir / "manual_test_report.xlsx", detail_rows, file_rows, summary)
    build_markdown(out_dir / "manual_test_report.md", summary, detail_rows, file_rows, format_stats, top_errors)
    build_improvement_plan(out_dir / "improvement_plan.md", top_errors, format_stats)

    (out_dir / "manual_test_results_raw.json").write_text(
        json.dumps({"summary": summary, "files": file_rows, "results": raw_results}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(json.dumps(summary, ensure_ascii=False, indent=2), flush=True)


if __name__ == "__main__":
    main()
