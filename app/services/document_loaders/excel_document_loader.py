from __future__ import annotations

import logging
import re
from pathlib import Path

from langchain_core.documents import Document
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.services.interfaces.document_loader import IDocumentLoader


logger = logging.getLogger(__name__)


class ExcelDocumentLoader(IDocumentLoader):
    _UNIT_RE = re.compile(r"^(.+?)\s*[\[(]([^\])]+)[\])]$")

    def supports(self, file_extension: str) -> bool:
        return file_extension.lower() in (".xlsx", ".xls", ".xlsm")

    def load(self, file_path: Path) -> list[Document]:
        try:
            documents = self._load_with_parser(
                file_path,
                parser_name="primary",
                read_only=False,
                data_only=False,
            )
            logger.info(
                "excel_loader_primary_success file=%s documents=%s",
                file_path.name,
                len(documents),
            )
            return documents
        except Exception:
            logger.warning(
                "excel_loader_primary_failed file=%s fallback=readonly_data_only",
                file_path.name,
                exc_info=True,
            )

        try:
            documents = self._load_with_parser(
                file_path,
                parser_name="fallback_readonly",
                read_only=True,
                data_only=True,
            )
            logger.info(
                "excel_loader_fallback_success file=%s documents=%s",
                file_path.name,
                len(documents),
            )
            return documents
        except Exception as exc:
            logger.error(
                "excel_loader_fallback_failed file=%s",
                file_path.name,
                exc_info=True,
            )
            raise RuntimeError(f"Failed to parse spreadsheet {file_path.name}") from exc

    def _load_with_parser(
        self,
        file_path: Path,
        *,
        parser_name: str,
        read_only: bool,
        data_only: bool,
    ) -> list[Document]:
        workbook = load_workbook(
            str(file_path),
            read_only=read_only,
            data_only=data_only,
        )
        value_workbook = None
        if not data_only:
            try:
                value_workbook = load_workbook(
                    str(file_path),
                    read_only=read_only,
                    data_only=True,
                )
            except Exception:
                logger.warning(
                    "excel_loader_cached_values_unavailable file=%s parser=%s",
                    file_path.name,
                    parser_name,
                    exc_info=True,
                )
        documents: list[Document] = []

        try:
            total_data_cells = 0
            total_tables = 0
            parsed_sheets = 0

            for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
                worksheet = workbook[sheet_name]
                value_worksheet = None
                if value_workbook is not None and sheet_name in value_workbook.sheetnames:
                    value_worksheet = value_workbook[sheet_name]
                try:
                    sheet_documents, sheet_data_cells, sheet_table_count = self._extract_sheet_documents(
                        file_path=file_path,
                        worksheet=worksheet,
                        value_worksheet=value_worksheet,
                        sheet_index=sheet_index,
                    )
                except Exception:
                    logger.warning(
                        "excel_loader_sheet_parse_failed file=%s parser=%s sheet=%s sheet_index=%s",
                        file_path.name,
                        parser_name,
                        sheet_name,
                        sheet_index,
                        exc_info=True,
                    )
                    continue

                parsed_sheets += 1
                total_data_cells += sheet_data_cells
                total_tables += sheet_table_count
                documents.extend(sheet_documents)

            if not documents:
                raise ValueError("No extractable spreadsheet content")

            logger.info(
                "excel_loader_summary file=%s parser=%s sheets=%s data_cells=%s tables_or_ranges=%s documents=%s",
                file_path.name,
                parser_name,
                parsed_sheets,
                total_data_cells,
                total_tables,
                len(documents),
            )
            return documents
        finally:
            workbook.close()
            if value_workbook is not None:
                value_workbook.close()

    def _extract_sheet_documents(
        self,
        *,
        file_path: Path,
        worksheet: Worksheet,
        value_worksheet: Worksheet | None,
        sheet_index: int,
    ) -> tuple[list[Document], int, int]:
        min_row, max_row, min_col, max_col = self._resolve_sheet_bounds(worksheet)
        merged_ranges = self._extract_merged_ranges(worksheet)
        hidden_sheet = str(getattr(worksheet, "sheet_state", "visible")).lower() != "visible"

        row_payloads: list[dict[str, object]] = []
        value_grid: list[list[str]] = []
        data_cell_count = 0

        for row in worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            row_number = int(row[0].row) if row else 0
            row_cells: list[dict[str, object]] = []
            row_values: list[str] = []
            row_has_data = False

            for cell in row:
                value_text, formula_text = self._stringify_cell_value(cell)
                if formula_text and value_worksheet is not None:
                    cached_value_text = self._stringify_cached_cell_value(value_worksheet, cell.coordinate)
                    if cached_value_text:
                        value_text = cached_value_text
                comment_text = self._extract_comment_text(cell)
                hyperlink_text = self._extract_hyperlink(cell)
                merged_range = self._resolve_merged_range(cell.coordinate, merged_ranges)

                has_signal = bool(value_text or formula_text or comment_text or hyperlink_text)
                if has_signal:
                    row_has_data = True
                    data_cell_count += 1

                row_values.append(value_text)
                row_cells.append(
                    {
                        "address": cell.coordinate,
                        "column_letter": get_column_letter(int(cell.column)),
                        "column_index": int(cell.column),
                        "value": value_text,
                        "formula": formula_text,
                        "comment": comment_text,
                        "hyperlink": hyperlink_text,
                        "merged_range": merged_range,
                    }
                )

            if not row_has_data:
                continue

            value_grid.append(row_values)
            row_payloads.append(
                {
                    "row_number": row_number,
                    "row_cells": row_cells,
                    "row_values": row_values,
                }
            )

        if not row_payloads:
            summary_doc = Document(
                page_content=self._build_empty_sheet_content(
                    file_name=file_path.name,
                    sheet_name=worksheet.title,
                    sheet_index=sheet_index,
                    sheet_hidden=hidden_sheet,
                ),
                metadata={
                    "source": str(file_path),
                    "extension": file_path.suffix.lower(),
                    "content_type": "spreadsheet_sheet_summary",
                    "file_name": file_path.name,
                    "sheet": worksheet.title,
                    "sheet_name": worksheet.title,
                    "sheet_index": sheet_index,
                    "sheet_hidden": hidden_sheet,
                    "row_count": 0,
                    "column_count": 0,
                    "data_cell_count": 0,
                    "table_count": 0,
                    "section_title": worksheet.title,
                },
            )
            return [summary_doc], 0, 0

        header_row_index = self._detect_header_row_index(value_grid)
        header_row_number = (
            int(row_payloads[header_row_index]["row_number"])
            if header_row_index is not None
            else None
        )
        headers = self._build_column_names(
            value_grid,
            header_row_index=header_row_index,
            column_count=max_col - min_col + 1,
        )
        header_units = self._extract_header_units(headers)

        table_specs = self._extract_table_specs(
            worksheet=worksheet,
            fallback_range=f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
        )

        sheet_documents: list[Document] = []
        table_summaries: list[str] = []
        formulas_count = 0
        comments_count = 0
        hyperlinks_count = 0

        for table_index, table_spec in enumerate(table_specs, start=1):
            table_name = table_spec["name"]
            table_kind = table_spec["kind"]
            table_range = table_spec["range"]

            table_document, table_metrics = self._build_table_document(
                file_path=file_path,
                worksheet=worksheet,
                sheet_index=sheet_index,
                sheet_hidden=hidden_sheet,
                table_name=table_name,
                table_kind=table_kind,
                table_range=table_range,
                row_payloads=row_payloads,
                global_headers=headers,
                header_units=header_units,
                header_row_number=header_row_number,
                merged_ranges=merged_ranges,
            )

            if table_document is None:
                logger.warning(
                    "excel_loader_table_empty file=%s sheet=%s table=%s range=%s",
                    file_path.name,
                    worksheet.title,
                    table_name,
                    table_range,
                )
                continue

            formulas_count += int(table_metrics.get("formula_cell_count", 0))
            comments_count += int(table_metrics.get("comment_count", 0))
            hyperlinks_count += int(table_metrics.get("hyperlink_count", 0))
            table_summaries.append(
                f"- {table_name} [{table_kind}] {table_range}: rows={table_metrics.get('row_count', 0)}"
            )
            sheet_documents.append(table_document)

            logger.info(
                "excel_loader_table_parsed file=%s sheet=%s table=%s range=%s rows=%s columns=%s formulas=%s comments=%s hyperlinks=%s",
                file_path.name,
                worksheet.title,
                table_name,
                table_range,
                table_metrics.get("row_count", 0),
                table_metrics.get("column_count", 0),
                table_metrics.get("formula_cell_count", 0),
                table_metrics.get("comment_count", 0),
                table_metrics.get("hyperlink_count", 0),
            )

        summary_content = self._build_sheet_summary_content(
            file_name=file_path.name,
            sheet_name=worksheet.title,
            sheet_index=sheet_index,
            sheet_hidden=hidden_sheet,
            headers=headers,
            header_units=header_units,
            row_payloads=row_payloads,
            table_summaries=table_summaries,
        )
        summary_doc = Document(
            page_content=summary_content,
            metadata={
                "source": str(file_path),
                "extension": file_path.suffix.lower(),
                "content_type": "spreadsheet_sheet_summary",
                "file_name": file_path.name,
                "sheet": worksheet.title,
                "sheet_name": worksheet.title,
                "sheet_index": sheet_index,
                "sheet_hidden": hidden_sheet,
                "row_count": len(row_payloads),
                "column_count": len(headers),
                "headers": headers,
                "header_units": header_units,
                "has_header": header_row_number is not None,
                "header_row_number": header_row_number,
                "data_cell_count": data_cell_count,
                "table_count": len(table_summaries),
                "formula_cell_count": formulas_count,
                "comment_count": comments_count,
                "hyperlink_count": hyperlinks_count,
                "merged_ranges": merged_ranges[:128],
                "section_title": worksheet.title,
            },
        )
        sheet_documents.insert(0, summary_doc)

        return sheet_documents, data_cell_count, len(table_summaries)

    def _build_table_document(
        self,
        *,
        file_path: Path,
        worksheet: Worksheet,
        sheet_index: int,
        sheet_hidden: bool,
        table_name: str,
        table_kind: str,
        table_range: str,
        row_payloads: list[dict[str, object]],
        global_headers: list[str],
        header_units: dict[str, str],
        header_row_number: int | None,
        merged_ranges: list[str],
    ) -> tuple[Document | None, dict[str, int]]:
        min_col, min_row, max_col, max_row = range_boundaries(table_range)
        table_headers = global_headers[min_col - 1:max_col] if global_headers else []

        if not table_headers:
            table_headers = [f"column_{index}" for index in range(min_col, max_col + 1)]

        has_header = bool(header_row_number is not None and min_row <= header_row_number <= max_row)
        data_start_row = (header_row_number + 1) if has_header else min_row

        structured_rows: list[dict[str, object]] = []
        formula_cells = 0
        comment_cells = 0
        hyperlink_cells = 0

        for row_payload in row_payloads:
            row_number = int(row_payload["row_number"])
            if row_number < data_start_row or row_number > max_row:
                continue

            row_cells = list(row_payload["row_cells"])
            scoped_cells = row_cells[min_col - 1:max_col]

            value_map: dict[str, str] = {}
            scoped_cell_payloads: list[dict[str, str]] = []

            for index, cell_payload in enumerate(scoped_cells):
                header_name = table_headers[index] if index < len(table_headers) else f"column_{min_col + index}"
                value_text = str(cell_payload.get("value") or "").strip()
                formula_text = str(cell_payload.get("formula") or "").strip()
                comment_text = str(cell_payload.get("comment") or "").strip()
                hyperlink_text = str(cell_payload.get("hyperlink") or "").strip()

                if value_text:
                    value_map[header_name] = value_text

                if formula_text:
                    formula_cells += 1
                if comment_text:
                    comment_cells += 1
                if hyperlink_text:
                    hyperlink_cells += 1

                scoped_cell_payloads.append(
                    {
                        "address": str(cell_payload.get("address") or ""),
                        "header": header_name,
                        "value": value_text,
                        "formula": formula_text,
                        "comment": comment_text,
                        "hyperlink": hyperlink_text,
                        "merged_range": str(cell_payload.get("merged_range") or ""),
                    }
                )

            if not value_map and not any(
                cell.get("formula") or cell.get("comment") or cell.get("hyperlink")
                for cell in scoped_cell_payloads
            ):
                continue

            structured_rows.append(
                {
                    "row_number": row_number,
                    "row_range": f"{get_column_letter(min_col)}{row_number}:{get_column_letter(max_col)}{row_number}",
                    "values": value_map,
                    "cells": scoped_cell_payloads,
                }
            )

        if not structured_rows:
            return None, {}

        table_row_start = int(structured_rows[0]["row_number"])
        table_row_end = int(structured_rows[-1]["row_number"])

        content = self._build_table_content(
            file_name=file_path.name,
            sheet_name=worksheet.title,
            sheet_index=sheet_index,
            sheet_hidden=sheet_hidden,
            table_name=table_name,
            table_kind=table_kind,
            table_range=table_range,
            headers=table_headers,
            header_units=header_units,
            structured_rows=structured_rows,
        )

        metrics = {
            "row_count": len(structured_rows),
            "column_count": len(table_headers),
            "formula_cell_count": formula_cells,
            "comment_count": comment_cells,
            "hyperlink_count": hyperlink_cells,
        }

        document = Document(
            page_content=content,
            metadata={
                "source": str(file_path),
                "extension": file_path.suffix.lower(),
                "content_type": "spreadsheet_table",
                "file_name": file_path.name,
                "sheet": worksheet.title,
                "sheet_name": worksheet.title,
                "sheet_index": sheet_index,
                "sheet_hidden": sheet_hidden,
                "table_name": table_name,
                "table_kind": table_kind,
                "range_address": table_range,
                "row_range": f"{table_row_start}:{table_row_end}",
                "column_range": f"{get_column_letter(min_col)}:{get_column_letter(max_col)}",
                "has_header": has_header,
                "header_row_number": header_row_number,
                "headers": table_headers,
                "header_units": {
                    header: header_units.get(header, "")
                    for header in table_headers
                },
                "row_count": len(structured_rows),
                "column_count": len(table_headers),
                "formula_cell_count": formula_cells,
                "comment_count": comment_cells,
                "hyperlink_count": hyperlink_cells,
                "merged_ranges": merged_ranges[:128],
                "structured_rows": structured_rows,
                "section_title": f"{worksheet.title}:{table_name}",
            },
        )
        return document, metrics

    @staticmethod
    def _resolve_sheet_bounds(worksheet: Worksheet) -> tuple[int, int, int, int]:
        min_row = max(1, int(getattr(worksheet, "min_row", 1) or 1))
        max_row = max(min_row, int(getattr(worksheet, "max_row", min_row) or min_row))
        min_col = max(1, int(getattr(worksheet, "min_column", 1) or 1))
        max_col = max(min_col, int(getattr(worksheet, "max_column", min_col) or min_col))

        try:
            dimension = str(worksheet.calculate_dimension() or "").strip()
        except Exception:
            dimension = ""

        if dimension and ":" in dimension and dimension != "A1:A1":
            dim_min_col, dim_min_row, dim_max_col, dim_max_row = range_boundaries(dimension)
            min_row = min(min_row, dim_min_row)
            max_row = max(max_row, dim_max_row)
            min_col = min(min_col, dim_min_col)
            max_col = max(max_col, dim_max_col)

        return min_row, max_row, min_col, max_col

    @staticmethod
    def _extract_merged_ranges(worksheet: Worksheet) -> list[str]:
        merged = getattr(worksheet, "merged_cells", None)
        if merged is None:
            return []
        return [str(item) for item in getattr(merged, "ranges", [])]

    @staticmethod
    def _resolve_merged_range(cell_address: str, merged_ranges: list[str]) -> str:
        for merged_range in merged_ranges:
            try:
                min_col, min_row, max_col, max_row = range_boundaries(merged_range)
                col, row = range_boundaries(f"{cell_address}:{cell_address}")[0:2]
            except Exception:
                continue

            if min_col <= col <= max_col and min_row <= row <= max_row:
                return merged_range
        return ""

    @classmethod
    def _extract_table_specs(cls, *, worksheet: Worksheet, fallback_range: str) -> list[dict[str, str]]:
        specs: list[dict[str, str]] = []
        tables = getattr(worksheet, "tables", None)

        if tables:
            for table_name, table_ref in tables.items():
                try:
                    ref = str(getattr(table_ref, "ref", "") or "")
                except Exception:
                    ref = ""
                if not ref:
                    continue
                specs.append(
                    {
                        "name": str(table_name),
                        "kind": "named_table",
                        "range": ref,
                    }
                )

        if specs:
            return specs

        return [
            {
                "name": "used_range_1",
                "kind": "used_range",
                "range": fallback_range,
            }
        ]

    @staticmethod
    def _extract_comment_text(cell) -> str:
        comment = getattr(cell, "comment", None)
        if comment is None:
            return ""
        return str(getattr(comment, "text", "") or "").strip()

    @staticmethod
    def _extract_hyperlink(cell) -> str:
        hyperlink = getattr(cell, "hyperlink", None)
        if hyperlink is None:
            return ""
        target = str(getattr(hyperlink, "target", "") or "").strip()
        if target:
            return target
        return str(getattr(hyperlink, "location", "") or "").strip()

    @staticmethod
    def _stringify_cell_value(cell) -> tuple[str, str]:
        raw_value = getattr(cell, "value", None)
        if raw_value is None:
            return "", ""

        text = str(raw_value).strip()
        if not text:
            return "", ""

        if isinstance(raw_value, str) and text.startswith("="):
            return "", text

        return text, ""

    @classmethod
    def _stringify_cached_cell_value(cls, worksheet: Worksheet, cell_coordinate: str) -> str:
        try:
            cell = worksheet[cell_coordinate]
        except Exception:
            return ""

        value_text, formula_text = cls._stringify_cell_value(cell)
        if formula_text:
            return ""
        return value_text

    @classmethod
    def _build_column_names(
        cls,
        raw_rows: list[list[str]],
        *,
        header_row_index: int | None,
        column_count: int,
    ) -> list[str]:
        if column_count <= 0:
            return []

        header_row = raw_rows[header_row_index] if header_row_index is not None and raw_rows else []
        names: list[str] = []
        seen: dict[str, int] = {}

        for column_index in range(column_count):
            raw_name = ""
            if column_index < len(header_row):
                raw_name = str(header_row[column_index]).strip()

            base_name = raw_name or f"column_{column_index + 1}"
            key = base_name.casefold()
            duplicate_count = seen.get(key, 0)
            seen[key] = duplicate_count + 1

            if duplicate_count > 0:
                base_name = f"{base_name}_{duplicate_count + 1}"

            names.append(base_name)

        return names

    @classmethod
    def _extract_header_units(cls, headers: list[str]) -> dict[str, str]:
        units: dict[str, str] = {}
        for header in headers:
            match = cls._UNIT_RE.match(str(header).strip())
            if match is None:
                units[header] = ""
                continue
            units[header] = str(match.group(2)).strip()
        return units

    @staticmethod
    def _looks_like_header(header: list[str]) -> bool:
        if not header:
            return False

        non_empty = [cell for cell in header if str(cell).strip()]
        if not non_empty:
            return False

        numeric_like = 0
        for cell in non_empty:
            normalized = str(cell).strip()
            if re.fullmatch(r"[-+]?\d+(?:[.,]\d+)?", normalized):
                numeric_like += 1

        return numeric_like < len(non_empty)

    @classmethod
    def _detect_header_row_index(cls, raw_rows: list[list[str]]) -> int | None:
        if not raw_rows:
            return None

        first_row_non_empty = cls._count_non_empty_cells(raw_rows[0])
        if len(raw_rows) > 1:
            second_row_non_empty = cls._count_non_empty_cells(raw_rows[1])
            if (
                first_row_non_empty <= 1
                and second_row_non_empty >= 2
                and cls._looks_like_header(raw_rows[1])
            ):
                return 1

        if cls._looks_like_header(raw_rows[0]):
            return 0

        for index, row in enumerate(raw_rows[1:5], start=1):
            if cls._count_non_empty_cells(row) < 2:
                continue
            if cls._looks_like_header(row):
                return index

        return None

    @staticmethod
    def _count_non_empty_cells(row: list[str]) -> int:
        return sum(1 for cell in row if str(cell).strip())

    @staticmethod
    def _build_empty_sheet_content(
        *,
        file_name: str,
        sheet_name: str,
        sheet_index: int,
        sheet_hidden: bool,
    ) -> str:
        return "\n".join(
            [
                f"File: {file_name}",
                f"Sheet: {sheet_name}",
                f"Sheet Index: {sheet_index}",
                f"Hidden Sheet: {sheet_hidden}",
                "Status: Empty sheet",
            ]
        ).strip()

    @staticmethod
    def _build_sheet_summary_content(
        *,
        file_name: str,
        sheet_name: str,
        sheet_index: int,
        sheet_hidden: bool,
        headers: list[str],
        header_units: dict[str, str],
        row_payloads: list[dict[str, object]],
        table_summaries: list[str],
    ) -> str:
        lines = [
            f"File: {file_name}",
            f"Sheet: {sheet_name}",
            f"Sheet Index: {sheet_index}",
            f"Hidden Sheet: {sheet_hidden}",
            f"Header Columns: {', '.join(headers) if headers else '(none)'}",
            f"Header Units: {', '.join(f'{key}={value}' for key, value in header_units.items() if value) or '(none)'}",
            f"Rows With Data: {len(row_payloads)}",
            f"Tables/Ranges: {len(table_summaries)}",
        ]

        if table_summaries:
            lines.append("Detected Tables/Ranges:")
            lines.extend(table_summaries[:24])

        preview_rows = row_payloads[:6]
        if preview_rows:
            lines.append("Row Preview:")
            for row_payload in preview_rows:
                row_number = int(row_payload["row_number"])
                row_values = [str(value).strip() for value in row_payload["row_values"] if str(value).strip()]
                compact = "; ".join(row_values[:8])
                if compact:
                    lines.append(f"- Row {row_number}: {compact}")

        return "\n".join(lines).strip()

    @staticmethod
    def _build_table_content(
        *,
        file_name: str,
        sheet_name: str,
        sheet_index: int,
        sheet_hidden: bool,
        table_name: str,
        table_kind: str,
        table_range: str,
        headers: list[str],
        header_units: dict[str, str],
        structured_rows: list[dict[str, object]],
    ) -> str:
        lines = [
            f"File: {file_name}",
            f"Sheet: {sheet_name}",
            f"Sheet Index: {sheet_index}",
            f"Hidden Sheet: {sheet_hidden}",
            f"Table: {table_name}",
            f"Table Type: {table_kind}",
            f"Range: {table_range}",
            f"Headers: {', '.join(headers) if headers else '(none)'}",
            f"Header Units: {', '.join(f'{header}={header_units.get(header, "")}' for header in headers if header_units.get(header, "")) or '(none)'}",
            f"Rows: {len(structured_rows)}",
            "Rows Preview:",
        ]

        for row_payload in structured_rows[:10]:
            row_number = int(row_payload.get("row_number", 0) or 0)
            row_range = str(row_payload.get("row_range") or "")
            value_map = row_payload.get("values") or {}
            if not isinstance(value_map, dict):
                continue

            cell_fragments: list[str] = []
            for key, value in list(value_map.items())[:12]:
                value_text = str(value).strip()
                if not value_text:
                    continue
                cell_fragments.append(f"{key}: {value_text}")

            if not cell_fragments:
                continue
            lines.append(f"- Row {row_number} [{row_range}]: {'; '.join(cell_fragments)}")

        return "\n".join(lines).strip()
